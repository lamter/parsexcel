#coding:utf-8
'''
Created on 2014-12-15

解析excel 的 *.xlsm 文件

@author: Shawn
'''


from openpyxl import load_workbook
wb = load_workbook(filename='../excelfile/officer.xlsm')
sheet_ranges = wb['officer']
print(sheet_ranges['A1'].value)



class XlsmWorkBook(object):
    """
    将 xlsm 文件中的一个表解析成 [{表头1:数值1, 表头2:数值2, ... }, ...]
    """
    def __init__(self, filename):
        """
        :param path: 指定一个路径
        :return:
        """

        self.wb = load_workbook(filename=filename)

        ''' 获得表名sheetName列表 '''
        # self.sheetNames =

        # ''' 获得表头列表 '''
        # self.headers = self.getHeaders()



    def getHeaders(self):
        '''
        定义表头的获取规则
        定义第一行为表头，出现空行后则不视为表头
        :return:
        '''





